from __future__ import unicode_literals
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from win32com.client import Dispatch

filename = 'C:/Users/LB/Desktop/Python/test.xlsx'
# 代码很多漏洞可以继续完善，由于时间问题暂时摆手
# 考勤表映射关系
#
# 1. 翻转字典
#   new_dict = {v : k for k, v in dict.items()}
# 
# 2. 函数：根据value，获取key
# def get_key (dict, value):
#   return [k for k, v in dict.items() if v == value]
#
# 3. 列表：[创建连个列表，利用一一对应的关系获取key对应的下标]
#  list(student.keys())
#  list (student.values()).index ('1004')
#  list (student.keys()) [list (student.values()).index ('1004')]
#
SheetMap0 = {
    'A':'B', #1. Sheet1 姓名 == Sheet2 姓名 [0A] - [1B]
    'X':'F', #2. Sheet1 加班-审批单统计 == Sheet2 加班审批时长 [23X] - [5F]
    'U':'L', #3. Sheet1 年假 == Sheet2 年假 [20U] - [11L]
    'S':'M', #4. Sheet1 事假 == Sheet2 事假 [18S] - [12M]
    'V':'N', #5.  Sheet1 病假 == Sheet2 病假 [21V] - [13N]
    'F':'C', #6. Sheet1 出勤天数 == Sheet2 实出勤天数 [5F] - [2C]
    'W':'O', #7. Sheet1 调休 == Sheet2 调休 [22W] - [14O]
    'T':'R', #8. Sheet1 婚假 == Sheet2 婚假 [19T] - [17R]
    'R':'S', #9. Sheet1 丧假 == Sheet2 丧假 [17R] - [18S]
    'G':'E', #10.  Sheet1 工作时长G == Sheet2 月总工时E 【时间单位转换-代码或者用公式计算】
}

# 考勤表映射关系
SheetMap = {
    'B':'A', #1. Sheet1 姓名 == Sheet2 姓名 [0A] - [1B]
    'F':'X', #2. Sheet1 加班-审批单统计 == Sheet2 加班审批时长 [23X] - [5F]
    'L':'U', #3. Sheet1 年假 == Sheet2 年假 [20U] - [11L]
    'M':'S', #4. Sheet1 事假 == Sheet2 事假 [18S] - [12M]
    'N':'V', #5.  Sheet1 病假 == Sheet2 病假 [21V] - [13N]
    'C':'F', #6. Sheet1 出勤天数 == Sheet2 实出勤天数 [5F] - [2C]
    'O':'W', #7. Sheet1 调休 == Sheet2 调休 [22W] - [14O]
    'R':'T', #8. Sheet1 婚假 == Sheet2 婚假 [19T] - [17R]
    'S':'R', #9. Sheet1 丧假 == Sheet2 丧假 [17R] - [18S]
    'E':'G', #9. Sheet1 丧假 == Sheet2 丧假 [17R] - [18S]
}

# 需要处理的数据
data = {}

def just_open(filename):
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(filename)
    xlBook.Save()
    xlBook.Close()

# Sheet1 工作时长G == Sheet2 月总工时E 【时间单位转换-代码或者用公式计算】
def timeTranslate(timeStr):
    res = 0
    timeStr = timeStr.replace('小时',',')
    timeStr = timeStr.replace('分钟',',')
    timeList = timeStr.split(',')
    if len(timeList)>=2 :
        res = float(timeList[0]) + float(timeList[1])/60
    print(timeList)
    return round(res,1)

# 读指定列的数据
def readExcelColumns(filename, sheetName, targetColums):
    # 读xlsm文件，data_only=True
    # keep_vba=True是因为我的xlsm文件里面有vb脚本，如果不加上这个，会报错
    workbook_ = load_workbook(filename,keep_vba=True,data_only=True)
    sheet = workbook_.get_sheet_by_name(sheetName)
    for columindex in targetColums:
        columdata = []
        for cell in sheet[columindex]:
            if cell.row >= 2:
                columdata.append(cell.value)
        data[SheetMap0[columindex]]=columdata

def writeExcelColums(filename, sheetName, targetColums, data):
    workbook_ = load_workbook(filename,keep_vba=True,data_only=False)
    sheet = workbook_.get_sheet_by_name(sheetName)
    for columindex in targetColums :
        for cell in sheet[columindex] :
            if (cell.row > 2 and cell.row < len(data[columindex])) :
                if (columindex!='E') :
                    cell.value=data[columindex][cell.row-2]
                else :
                    #203小时24分钟
                    cell.value=timeTranslate(data[columindex][cell.row-2]) 
    workbook_.save(filename)  # 保存文件

readExcelColumns(filename, "Sheet1", list(SheetMap0.keys()))
writeExcelColums(filename, "Sheet2", list(SheetMap0.values()), data)

# 利用win32 api打开这个xlsm文件，保存后再关闭（其实就跟人工打开文件再保存关闭一个道理）
just_open(filename)

